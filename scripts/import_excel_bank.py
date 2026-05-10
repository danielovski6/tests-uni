from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_EXCEL = Path(r"C:\Users\Daniel\Downloads\preguntas_tipo_test_fisioterapia_tema_1.xlsx")
OUT_PATH = Path("data/question-bank.json")


def slug(value: str) -> str:
    value = value.lower().strip()
    replacements = {
        "á": "a",
        "é": "e",
        "í": "i",
        "ó": "o",
        "ú": "u",
        "ü": "u",
        "ñ": "n",
    }
    for src, dst in replacements.items():
        value = value.replace(src, dst)
    return re.sub(r"[^a-z0-9]+", "-", value).strip("-") or "general"


def infer_topic(path: Path) -> tuple[str, str]:
    match = re.search(r"tema[_ -]*(\d+)", path.stem, re.IGNORECASE)
    if match:
        number = match.group(1)
        return f"tema-{number}", f"Tema {number}"
    return "general", "General"


def norm(value: str) -> str:
    return slug(value).replace("-", "_")


def row_get(row: tuple, index: dict[str, int], *names: str) -> str:
    for name in names:
        pos = index.get(norm(name))
        if pos is not None:
            return str(row[pos] or "").strip()
    return ""


def normalize_difficulty(value: str) -> str:
    normalized = slug(value)
    if normalized in {"facil", "easy"}:
        return "easy"
    if normalized in {"media", "medio", "medium", "normal"}:
        return "medium"
    if normalized in {"dificil", "hard"}:
        return "hard"
    return "medium"


def import_excel(path: Path) -> dict:
    wb = load_workbook(path, data_only=True)
    if "Preguntas_detalle" in wb.sheetnames:
        ws = wb["Preguntas_detalle"]
    elif "preguntas" in wb.sheetnames:
        ws = wb["preguntas"]
    else:
        ws = wb.active

    headers = [str(cell.value or "").strip() for cell in ws[1]]
    index = {norm(name): pos for pos, name in enumerate(headers)}
    required_groups = [
        ("Asignatura", "asignatura"),
        ("Pregunta", "pregunta"),
        ("Respuesta A", "respuesta_a"),
        ("Respuesta B", "respuesta_b"),
        ("Respuesta C", "respuesta_c"),
        ("Respuesta D", "respuesta_d"),
        ("Respuesta correcta", "respuesta_correcta"),
    ]
    missing = [names[0] for names in required_groups if not any(norm(name) in index for name in names)]
    if missing:
        raise ValueError(f"Missing columns: {', '.join(missing)}")

    fallback_topic_id, fallback_topic_name = infer_topic(path)
    subjects: dict[str, dict] = {}
    answer_letters = {"A": 0, "B": 1, "C": 2, "D": 3}

    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        subject_name = row_get(row, index, "Asignatura", "asignatura") or "General"
        topic_name = row_get(row, index, "Tema", "tema") or fallback_topic_name
        topic_id = slug(topic_name)
        prompt = row_get(row, index, "Pregunta", "pregunta")
        options = [
            row_get(row, index, "Respuesta A", "respuesta_a"),
            row_get(row, index, "Respuesta B", "respuesta_b"),
            row_get(row, index, "Respuesta C", "respuesta_c"),
            row_get(row, index, "Respuesta D", "respuesta_d"),
        ]
        correct_letter = row_get(row, index, "Respuesta correcta", "respuesta_correcta").upper()
        difficulty = normalize_difficulty(row_get(row, index, "Dificultad", "dificultad"))
        explanation = (
            row_get(row, index, "Explicacion respuesta correcta", "explicacion_respuesta_correcta")
            or row_get(row, index, "Texto respuesta correcta", "texto_respuesta_correcta")
        )
        if not prompt or any(not option for option in options) or correct_letter not in answer_letters:
            continue

        subject_id = slug(subject_name)
        if subject_id not in subjects:
            subjects[subject_id] = {
                "id": subject_id,
                "name": subject_name,
                "topics": [],
                "questions": [],
            }
        if not any(topic["id"] == topic_id for topic in subjects[subject_id]["topics"]):
            subjects[subject_id]["topics"].append({"id": topic_id, "name": topic_name})

        question_number = len(subjects[subject_id]["questions"]) + 1
        correct_index = answer_letters[correct_letter]
        subjects[subject_id]["questions"].append(
            {
                "id": f"{subject_id}-{topic_id}-{question_number:03}",
                "topicId": topic_id,
                "prompt": prompt,
                "options": options,
                "correctIndex": correct_index,
                "difficulty": difficulty,
                "explanation": explanation or f"Respuesta correcta: {options[correct_index]}",
                "source": f"{path.name}, fila {row_number}",
            }
        )

    return {"version": "2026-05-10-excel", "subjects": list(subjects.values())}


def merge_banks(banks: list[dict]) -> dict:
    merged = {"version": "2026-05-10-excel", "subjects": []}
    subjects: dict[str, dict] = {}
    for bank in banks:
        for incoming in bank["subjects"]:
            subject = subjects.setdefault(
                incoming["id"],
                {"id": incoming["id"], "name": incoming["name"], "topics": [], "questions": []},
            )
            for topic in incoming["topics"]:
                if not any(existing["id"] == topic["id"] for existing in subject["topics"]):
                    subject["topics"].append(topic)
            existing_ids = {question["id"] for question in subject["questions"]}
            for question in incoming["questions"]:
                new_question = dict(question)
                if new_question["id"] in existing_ids:
                    new_question["id"] = f"{new_question['id']}-{len(subject['questions']) + 1}"
                subject["questions"].append(new_question)
    merged["subjects"] = list(subjects.values())
    return merged


def main() -> None:
    excel_paths = [Path(arg) for arg in sys.argv[1:]] or [DEFAULT_EXCEL]
    bank = merge_banks([import_excel(path) for path in excel_paths])
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text(json.dumps(bank, ensure_ascii=False, indent=2), encoding="utf-8")
    total = sum(len(subject["questions"]) for subject in bank["subjects"])
    print(f"Wrote {total} questions from {len(excel_paths)} Excel file(s) to {OUT_PATH}")


if __name__ == "__main__":
    main()
